import json
import openpyxl
import read_name_list
import tkinter as tk
import datetime
import glob

class MakeGrade:
    def __init__(self, year, month, day, hour, minute):
        self.name_dict = {}
        self.name_list = [{i:0 for i in 'ABCDE'} for j in range(0, 3)]
        self.time_dict = {}
        
        self.workbook = openpyxl.Workbook(True)
        
        self.timezone = datetime.timezone(datetime.timedelta(hours=9))
        
        self.start = datetime.datetime(year, month, day, hour, minute, tzinfo=self.timezone)
        self.start.astimezone(self.timezone)
        
        self.read_name_list()
        self.read_json()
        
        self.write_final_lank()
        self.write_class()
        self.save_excel()

    def read_json(self):
        length = 0
        for file in glob.glob('./dates/*.json'):
            with open(file, 'r') as f:
                d = json.loads(f.read())
                length = max(length, d['checkpoint'])
        
        for i in self.time_dict.keys():
            self.time_dict[i] = [None for i in range(length)]
        
        for file in glob.glob('./dates/*.json'):
            with open(file, 'r') as f:
                d = json.loads(f.read())
                for i in self.name_dict.keys():
                    if i in d:
                        c = datetime.datetime.fromtimestamp(d[i])
                        c = c.replace(tzinfo=self.timezone)
                        self.time_dict[i][d['checkpoint'] - 1] = c
        
    
    def read_name_list(self):
        for i in read_name_list.ReadNameList().open():
            self.name_list[i[0] - 1][i[1]] += 1
            n = f'{i[0]}{i[1]}{str(i[2]).zfill(2)}'
            self.name_dict[n] = i[3]
            self.time_dict[n] = []
    
    def write_final_lank(self):
        self.workbook.create_sheet('最終順位', 0)
        sheet = self.workbook['最終順位']
        li = []
        for a, b in self.time_dict.items():
            if b[-1] is not None:
                li.append((b[-1], b[-1] - self.start, a))
        li.sort()
        
        sheet.append(('年', '組', '番号', '名前', '到着時間', 'タイム', ))
        for i, (a, b, c) in enumerate(li):
            sheet.append([c[0], c[1], c[2:], self.name_dict[c], str(a)[11:16], str(b)[:4]])
    
    def write_class(self):
        for i in range(1, 4):
            for j in 'ABCDE':
                li = []
                for t in range(1, self.name_list[i - 1][j] + 1):
                    li.append(self.time_dict[f'{i}{j}{str(t).zfill(2)}'])
                if li:
                    self.workbook.create_sheet(f'{i}{j}', -1)
                    sheet = self.workbook[f'{i}{j}']
                    
                    sheet.append(['名前', *[f'第{a}関門' for a in range(1, len(li[0]) + 1)], '', *[f'{a}~{a+1}' for a in range(len(li[0]))]])
                    
                    for x, y in enumerate(li):
                        time = []
                        rap = []
                        before = self.start
                        for a in y:
                            if a is None:
                                time.append('記録なし')
                                rap.append('記録なし')
                                before = None
                            else:
                                time.append(str(a)[11:16])
                                if before is None:
                                    rap.append('記録なし')
                                else:
                                    rap.append(str(a - before)[:4])
                                before = a
                        sheet.append([self.name_dict[f'{i}{j}{str(x + 1).zfill(2)}']] + time + [''] + rap)

    def save_excel(self):
        self.workbook.save('結果.xlsx')


class AskDateAndTime(tk.Frame):
    def __init__(self, root):
        super().__init__(root)
        self.root = root
        self.write_screen()
    
    def write_screen(self):
        now = datetime.datetime.now()
        self.font = ('', 20)
        self.explanation_label = tk.Label(text='開始日時を指定してください。', font=self.font)
        
        self.year_val = tk.IntVar(value=now.year)
        self.year_spinbox = tk.Spinbox(from_=2000, to=9999, textvariable=self.year_val, width=4, font=self.font)
        self.year_label = tk.Label(text='年', font=self.font)
        
        self.month_val = tk.IntVar(value=now.month)
        self.month_spinbox = tk.Spinbox(from_=1, to=12, textvariable=self.month_val, width=2, font=self.font)
        self.month_label = tk.Label(text='月', font=self.font)
        
        self.day_val = tk.IntVar(value=now.day)
        self.day_spinbox = tk.Spinbox(from_=1, to=31, textvariable=self.day_val, width=2, font=self.font)
        self.day_label = tk.Label(text='日', font=self.font)
        
        self.hour_val = tk.IntVar(value=6)
        self.hour_spinbox = tk.Spinbox(from_=0, to=24, textvariable=self.hour_val, width=2, font=self.font)
        self.hour_label = tk.Label(text='時', font=self.font)

        self.minute_val = tk.IntVar(value=0)
        self.minute_spinbox = tk.Spinbox(from_=0, to=60, textvariable=self.minute_val, width=2, font=self.font)
        self.minute_label = tk.Label(text='分', font=self.font)
        
        self.button = tk.Button(text='決定', command=self.pressed, font=self.font)
        
        self.explanation_label.grid(column=0, row=0, columnspan=6)
        
        self.year_spinbox.grid(column=0, row=1)
        self.year_label.grid(column=1, row=1)
        self.month_spinbox.grid(column=2, row=1)
        self.month_label.grid(column=3, row=1)
        self.day_spinbox.grid(column=4, row=1)
        self.day_label.grid(column=5, row=1)
        
        self.hour_spinbox.grid(column=0, row=2)
        self.hour_label.grid(column=1, row=2)
        self.minute_spinbox.grid(column=4, row=2)
        self.minute_label.grid(column=5, row=2)
        
        self.button.grid(column=0, row=3, columnspan=6)
    
    def pressed(self):
        grade = MakeGrade(self.year_val.get(), self.month_val.get(), self.day_val.get(), self.hour_val.get(), self.minute_val.get())
        root.destroy()

if __name__ == '__main__':
    root = tk.Tk()
    root.title('Make Grade')
    ask_date = AskDateAndTime(root)
    ask_date.mainloop()